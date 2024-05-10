from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def loadSpreadSheet(sheet):
    wb = load_workbook(str(sheet))
    ws = wb.active
    return ws
 
def getHeaderNames(row, column, worksheet):
    infoCategories = []
    for col in worksheet.iter_cols(min_row=row, max_row=row, min_col=column):
        for cell in col:
            infoCategories.append(cell.value)
    return infoCategories

def getSlideDetails(categories, worksheet, row, column):
    slideDetails = []
    for row in worksheet.iter_rows(min_row=(row + 1), min_col=column, values_only=True):
        slideDetails.append(row)
    return slideDetails
